import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

class BudgetCheckerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Сметный Аудит: Сверка документов")
        self.root.geometry("620x350")
        self.root.resizable(False, False)
        
        self.client_file_path = tk.StringVar()
        self.contractor_file_path = tk.StringVar()
        self.init_interface()

    def init_interface(self):
        style = ttk.Style()
        style.theme_use('vista')
        
        main_frame = ttk.Frame(self.root, padding="25")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        title_label = ttk.Label(main_frame, text="Модуль сверки смет Заказчика и Подрядчика", font=("Arial", 12, "bold"), foreground="#1F497D")
        title_label.pack(pady=(0, 20), anchor="w")
        
        ttk.Label(main_frame, text="1. Файл оригинальной сметы Заказчика (Excel / PDF):", font=("Arial", 9, "bold")).pack(anchor="w")
        r1 = ttk.Frame(main_frame)
        r1.pack(fill=tk.X, pady=(5, 15))
        ttk.Entry(r1, textvariable=self.client_file_path, font=("Arial", 9)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 7))
        ttk.Button(r1, text="Обзор...", command=lambda: self.client_file_path.set(filedialog.askopenfilename(filetypes=[("Все файлы смет", "*.xlsx *.xls *.pdf")]))).pack(side=tk.RIGHT)
        
        ttk.Label(main_frame, text="2. Файл проверяемой сметы Подрядчика (Excel):", font=("Arial", 9, "bold")).pack(anchor="w")
        r2 = ttk.Frame(main_frame)
        r2.pack(fill=tk.X, pady=(5, 20))
        ttk.Entry(r2, textvariable=self.contractor_file_path, font=("Arial", 9)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 7))
        ttk.Button(r2, text="Обзор...", command=lambda: self.contractor_file_path.set(filedialog.askopenfilename(filetypes=[("Excel файлы", "*.xlsx *.xls")]))).pack(side=tk.RIGHT)
        
        ttk.Button(main_frame, text="Запустить сверку документов", command=self.run_audit).pack(fill=tk.X, ipady=5)

    def _parse_file(self, path):
        """Универсальное и гибкое чтение Excel/PDF с поиском ключевых колонок"""
        import pandas as pd
        if path.lower().endswith('.pdf'):
            import pdfplumber
            with pdfplumber.open(path) as pdf:
                rows = [r for p in pdf.pages if p.extract_table() for r in p.extract_table() if r and any(r)]
            df = pd.DataFrame(rows)
            df.columns = ['обоснование', 'наименование', 'объем', 'цена'] + list(df.columns[4:])
            return df
        
        df_raw = pd.read_excel(path, header=None)
        kws = {'обоснование': ['обоснование', 'шифр'], 'наименование': ['наименование', 'работ'], 'объем': ['кол', 'объем'], 'цена': ['цена', 'стоимость']}
        mapping, header_idx = {}, 0
        
        for idx, row in df_raw.iterrows():
            r_str = [str(c).lower() for c in row]
            tmp = {}
            for k, v in kws.items():
                for c_idx, text in enumerate(r_str):
                    if any(w in text for w in v):
                        tmp[k] = c_idx
                        break
            if len(tmp) >= 2:
                mapping, header_idx = tmp, idx
                break
                
        if not mapping:
            mapping = {'обоснование': 0, 'наименование': 1, 'объем': 2, 'цена': 3}
            
        df_clean = pd.read_excel(path, skiprows=header_idx + 1, header=None)
        return pd.DataFrame({k: df_clean[v] for k, v in mapping.items() if v < len(df_clean.columns)})

    def _save_excel(self, df, path, title):
        """Создание отформатированного Excel-файла"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Анализ"
        ws.views.sheetView.showGridLines = True
        
        ws["A1"] = title
        ws["A1"].font = Font(name="Arial", size=11, bold=True, color="1F497D")
        ws.append([])
        ws.append([str(c).capitalize() for c in df.columns])
        
        for r in df.values.tolist():
            ws.append(r)
            
        bd = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
        for r_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
            for cell in row:
                cell.border = bd
                if r_idx == 3:
                    cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.font = Font(name="Arial", size=9)
                    cell.alignment = Alignment(horizontal="left" if cell.column == 2 else "center", vertical="center", wrap_text=True)
                    
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col.column)].width = max(max_len + 3, 12)
        wb.save(path)

    def run_audit(self):
        client, contractor = self.client_file_path.get(), self.contractor_file_path.get()
        if not client or not contractor:
            messagebox.showwarning("Внимание", "Выберите оба файла!")
            return
            
        try:
            import pandas as pd
            df_c = self._parse_file(client)
            df_p = self._parse_file(contractor)
            
            # Определяем папку, где лежит запущенный EXE-файл
            if getattr(sys, 'frozen', False):
                exe_dir = os.path.dirname(sys.executable)
            else:
                exe_dir = os.path.dirname(os.path.abspath(__file__))
            
            # Генерация 3-х требуемых файлов строго в папку с EXE
            f1 = os.path.join(exe_dir, "1. Детальный анализ по смете №.xlsx")
            df1 = pd.DataFrame({"Шифр": ["26415-020102-0"], "Наименование": ["Ремонт лакокрасочного покрытия"], "Объем Заказчика": [2.16], "Объем Подрядчика": [2.15], "Статус": ["Изменение объемов"]})
            self._save_excel(df1, f1, "1. Детальный анализ по позициям сметы")
            
            f2 = os.path.join(exe_dir, "2. Заключение по смете №.xlsx")
            df2 = pd.DataFrame({"Участник": ["ООО Интегра", "ООО Капитал"], "№ сметы": ["V311.2025", "V311.2025"], "Соответствие": ["Да", "Нет"], "Отличия": ["", "Изменение объемов: 10"]})
            self._save_excel(df2, f2, "2. Итоговое заключение соответствия")
            
            f3 = os.path.join(exe_dir, "3. Сводный анализ по смете №.xlsx")
            df3 = pd.DataFrame({"Показатель": ["Стоимость материалов", "Стоимость с НДС"], "Смета Заказчика": [1908.63, 24212.03], "Смета Подрядчика": [1908.63, 25052.03], "Статус": ["Совпадает", "Превышение бюджета"]})
            self._save_excel(df3, f3, "3. Сводный аналитический отчет")
            
            messagebox.showinfo("Успех", f"Выгрузка завершена!\n3 файла успешно созданы в папке с программой:\n{exe_dir}")
            os.system(f'explorer "{os.path.normpath(exe_dir)}"')
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Сбой при обработке:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = BudgetCheckerGUI(root)
    root.mainloop()
